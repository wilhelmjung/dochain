"""Excel exporter — converts structured OCR results to Excel spreadsheet."""

from __future__ import annotations

import re
from dataclasses import dataclass, field, fields
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# -----------------------------------------------------------------------
# Invoice type mapping
# -----------------------------------------------------------------------
# Baidu OCR InvoiceType / InvoiceTypeOrg → standardised 票种
_INVOICE_TYPE_MAP: dict[str, str] = {
    # 铁路电子客票 (from train_ticket mode)
    "train_ticket": "数电发票（铁路电子客票）",
    # 增值税专用发票
    "电子发票(专用发票)": "数电发票（增值税专用发票）",
    "电子发票(增值税专用发票)": "数电发票（增值税专用发票）",
    "增值税专用发票": "数电发票（增值税专用发票）",
    "专用发票": "数电发票（增值税专用发票）",
    # 普通发票
    "电子发票(普通发票)": "数电发票（普通发票）",
    "增值税普通发票": "数电发票（普通发票）",
    "普通发票": "数电发票（普通发票）",
    "全电发票(普通发票)": "数电发票（普通发票）",
    # 航空运输电子客票行程单
    "航空运输电子客票行程单": "数电发票（航空运输电子客票行程单）",
    # 通行费
    "通行费发票": "数电发票（通行费发票）",
    "通行费": "数电发票（通行费发票）",
    # 区块链发票
    "区块链发票": "区块链发票",
    "深圳电子普通发票": "区块链发票",
}


def _map_invoice_type(raw_type: str, raw_type_org: str, mode: str) -> str:
    """Map raw Baidu OCR invoice type to standardised 票种."""
    if mode == "train_ticket":
        return "数电发票（铁路电子客票）"

    # Try InvoiceTypeOrg first (more specific), then InvoiceType
    for src in (raw_type_org, raw_type):
        src_stripped = src.strip()
        if src_stripped in _INVOICE_TYPE_MAP:
            return _INVOICE_TYPE_MAP[src_stripped]

    # Fuzzy match
    for src in (raw_type_org, raw_type):
        lower = src.strip()
        if "专用" in lower:
            return "数电发票（增值税专用发票）"
        if "普通" in lower:
            return "数电发票（普通发票）"
        if "铁路" in lower or "火车" in lower:
            return "数电发票（铁路电子客票）"
        if "航空" in lower or "行程单" in lower:
            return "数电发票（航空运输电子客票行程单）"
        if "通行费" in lower:
            return "数电发票（通行费发票）"
        if "区块链" in lower:
            return "区块链发票"

    return raw_type or raw_type_org or "未知"


# -----------------------------------------------------------------------
# Unified invoice record
# -----------------------------------------------------------------------

@dataclass
class InvoiceRecord:
    """A single row in the output Excel spreadsheet."""

    序号: int = 0
    数电发票号码: str = ""
    发票代码: str = ""
    发票号码: str = ""
    开票日期: str = ""
    金额: str = ""
    票面税额: str = ""
    有效抵扣税额: str = ""
    购买方识别号: str = ""
    销售方纳税人名称: str = ""
    销售方纳税人识别号: str = ""
    发票来源: str = ""
    票种: str = ""
    货物或劳务名称: str = ""
    旅客姓名: str = ""
    出行日期: str = ""
    乘客姓名: str = ""
    乘机日期: str = ""
    座位类型: str = ""


# Column header list (used for Excel generation)
COLUMN_HEADERS: list[str] = [f.name for f in fields(InvoiceRecord)]


def _get_word(obj) -> str:
    """Extract 'word' from a Baidu API dict item, or convert to str."""
    if isinstance(obj, dict):
        return obj.get("word", "")
    return str(obj) if obj else ""


def _join_commodity_names(names_list: list) -> str:
    """Join commodity names from a list of dict/str items."""
    if not names_list:
        return ""
    parts = [_get_word(n) for n in names_list]
    return "; ".join(p for p in parts if p)


def _clean_price(raw: str) -> str:
    """Extract numeric price from strings like '￥45.00元'."""
    if not raw:
        return ""
    match = re.search(r"[\d.]+", raw)
    return match.group() if match else raw


def _join_passenger_fields(passenger_list: list) -> str:
    """Join passenger-related list fields."""
    if not passenger_list:
        return ""
    parts = [_get_word(p) for p in passenger_list]
    return "; ".join(p for p in parts if p)


# -----------------------------------------------------------------------
# Extract record from structured OCR data
# -----------------------------------------------------------------------

def extract_invoice_record(
    mode: str,
    raw_data: dict,
    source_file: str = "",
    seq: int = 0,
) -> InvoiceRecord:
    """Convert raw Baidu OCR JSON to an InvoiceRecord.

    Args:
        mode: OCR mode that produced the result ("invoice", "train_ticket", "general").
        raw_data: The raw ``words_result`` dict (or full API response).
        source_file: Original filename (for 发票来源 column).
        seq: Sequence number.

    Returns:
        An InvoiceRecord with extracted fields.
    """

    # If raw_data is the full API response, get the words_result
    results = raw_data.get("words_result", raw_data)
    if not isinstance(results, dict):
        results = {}

    rec = InvoiceRecord(序号=seq, 发票来源=source_file)

    if mode == "train_ticket":
        rec.票种 = "数电发票（铁路电子客票）"
        rec.发票号码 = results.get("invoice_num", "") or results.get("ticket_num", "")
        rec.数电发票号码 = results.get("invoice_num", "")

        # 开票日期: prefer invoice_date, fallback to date
        rec.开票日期 = results.get("invoice_date", "") or results.get("date", "")

        # 金额: prefer ticket_rates, but some responses use ticket_price or fare
        raw_price = (
            results.get("ticket_rates", "")
            or results.get("ticket_price", "")
            or results.get("fare", "")
        )
        rec.金额 = _clean_price(raw_price)

        # 税额
        rec.票面税额 = results.get("tax", "")
        rec.有效抵扣税额 = rec.票面税额

        # 旅客
        rec.旅客姓名 = results.get("name", "")
        rec.出行日期 = results.get("date", "")
        rec.座位类型 = results.get("seat_category", "")

        # 货物或劳务名称: 出发→到达
        dep = results.get("starting_station", "")
        arr = results.get("destination_station", "")
        train_num = results.get("train_num", "")
        parts = []
        if dep and arr:
            parts.append(f"{dep}→{arr}")
        if train_num:
            parts.append(train_num)
        rec.货物或劳务名称 = " ".join(parts)

    elif mode == "invoice":
        # Standard invoice (VAT invoice / blockchain invoice / etc.)
        invoice_type = results.get("InvoiceType", "")
        invoice_type_org = results.get("InvoiceTypeOrg", "")
        rec.票种 = _map_invoice_type(invoice_type, invoice_type_org, mode)

        rec.发票号码 = results.get("InvoiceNum", "")
        rec.发票代码 = results.get("InvoiceCode", "")

        # 数电发票号码: if the invoice number is 20+ digits, it's a 数电 number
        inv_num = rec.发票号码
        if len(inv_num) >= 20:
            rec.数电发票号码 = inv_num

        rec.开票日期 = results.get("InvoiceDate", "")
        rec.金额 = results.get("TotalAmount", "")
        rec.票面税额 = results.get("TotalTax", "")
        rec.有效抵扣税额 = rec.票面税额  # default: same as 票面税额

        rec.购买方识别号 = results.get("PurchaserRegisterNum", "")
        rec.销售方纳税人名称 = results.get("SellerName", "")
        rec.销售方纳税人识别号 = results.get("SellerRegisterNum", "")

        # 货物或劳务名称
        rec.货物或劳务名称 = _join_commodity_names(results.get("CommodityName", []))

        # 旅客信息 (for 客票类)
        rec.旅客姓名 = _join_passenger_fields(results.get("PassengerName", []))
        rec.出行日期 = _join_passenger_fields(results.get("PassengerDate", []))
        rec.乘客姓名 = rec.旅客姓名  # same source for now
        rec.乘机日期 = rec.出行日期

        # 座位类型
        passenger_class = results.get("PassengerClass", [])
        if passenger_class:
            rec.座位类型 = _join_passenger_fields(passenger_class)

    else:
        # general mode — limited structured data available
        rec.票种 = "未知"

    return rec


# -----------------------------------------------------------------------
# Excel export
# -----------------------------------------------------------------------

def export_to_excel(records: list[InvoiceRecord], output_path: str | Path) -> Path:
    """Write a list of InvoiceRecords to an Excel file.

    Args:
        records: List of InvoiceRecord objects.
        output_path: Path to the .xlsx file to create.

    Returns:
        The Path of the written file.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"

    # ----- Styles -----
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ----- Write header row -----
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ----- Write data rows -----
    # Number columns that should be centered
    center_cols = {"序号"}
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, f in enumerate(fields(InvoiceRecord), start=1):
            value = getattr(rec, f.name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if f.name in center_cols:
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment

    # ----- Set column widths -----
    col_widths = {
        "序号": 6,
        "数电发票号码": 26,
        "发票代码": 16,
        "发票号码": 26,
        "开票日期": 18,
        "金额": 14,
        "票面税额": 12,
        "有效抵扣税额": 12,
        "购买方识别号": 24,
        "销售方纳税人名称": 28,
        "销售方纳税人识别号": 24,
        "发票来源": 24,
        "票种": 30,
        "货物或劳务名称": 34,
        "旅客姓名": 12,
        "出行日期": 18,
        "乘客姓名": 12,
        "乘机日期": 18,
        "座位类型": 10,
    }
    for col_idx, header in enumerate(COLUMN_HEADERS, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(header, 16)

    # ----- Freeze header row -----
    ws.freeze_panes = "A2"

    # ----- Auto-filter -----
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(output_path))
    return output_path
